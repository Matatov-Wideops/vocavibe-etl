import time

import boto3
import os
from os.path import exists
import re

from google.cloud import storage, bigquery
from tqdm import tqdm
import pandas as pd
import tempfile
import numpy as np
from datetime import datetime, timedelta

from cryptography.fernet import Fernet
from typing import List
from openpyxl import Workbook
from openpyxl.styles import PatternFill  

from src.keys import Keys
from src.settings import Settings
from src.database import Bucket, Entity, OnMed, FeelOnOff, Qnnrs, UPDRS, MoCA, PDQ8, FOG, SDQ, WOQ, Registration, Update, ExtraCols, Exercise, SamplerQnnrs

from src.utils import list_to_csv, get_os, run_shell_command, run_windows_shell_command, hash_phone_number, standardize_phone_number, sampler_phone_to_name, get_s3
from src.patterns import extract_from_filename, Patterns
from src.logger import load_processed_files, save_processed_filekey, load_blacklist
from src.resolve import resolve, change_columns, resolve_sessions
from src.plotting import Color, rgb_to_hex

import matplotlib
matplotlib.use('Agg')  # Switch to non-interactive backend
import matplotlib.pyplot as plt
plt.ioff()  # Turn off interactive mode



def combine_yahav_ec2() -> None:
    pd_yahav = pd.read_csv(Settings.USERS_YAHAV_CSV, dtype=str, index_col=0)
    pd_yahav[ExtraCols.PASSWORD.value] = np.nan
    pd_ec2 = pd.read_csv(Settings.USERSPD, dtype=str)
    pd_ec2.columns = [ExtraCols.USER_PHONE.value, ExtraCols.PASSWORD.value]
    pd_ec2['username'] = pd_ec2[ExtraCols.USER_PHONE.value].apply(hash_phone_number)
    combo = pd.concat([pd_yahav, pd_ec2], ignore_index=True)

    phonesNpasswords = pd.read_csv('resources/passwords.csv', dtype=str)
    # Merge the dataframes on ExtraCols.USER_PHONE.value with an outer join to ensure all users are included
    combined_df = pd.merge(combo, phonesNpasswords, on=ExtraCols.USER_PHONE.value, how='outer', suffixes=('_users', '_passwords'))

    # Fill missing passwords in users_df with passwords from passwords_df
    combined_df[ExtraCols.PASSWORD.value] = combined_df['password_users'].combine_first(combined_df['password_passwords'])

    # Drop the now redundant columns
    combined_df.drop(columns=['password_users', 'password_passwords'], inplace=True)

    # to_drop = combined_df[combined_df['username'].isin(IgnoreUsers.PD)].index
    # combined_df = combined_df.drop(to_drop)
    
    # Sort the dataframe such that rows with passwords come first
    # combined_df = combined_df.sort_values(by=ExtraCols.PASSWORD, ascending=False)
    # Drop duplicates, keeping the first occurrence (which will have the password if available)
    combined_df = combined_df.drop_duplicates(subset=ExtraCols.USER_PHONE.value, keep='last', ignore_index=True)
    # combined_df = combined_df.reset_index(drop=True)
    combined_df.to_csv(Settings.USERS_EC2_CSV)



def healthy_ec2() -> None:
    hc_ec2 = pd.read_csv(Settings.USERSHC, dtype=str)
    hc_ec2.columns = [ExtraCols.USER_PHONE.value, ExtraCols.PASSWORD.value]
    hc_ec2[Bucket.USERNAME] = hc_ec2[ExtraCols.USER_PHONE.value].apply(hash_phone_number)
    hc_ec2[Bucket.USERNAME] = "hc_" + hc_ec2[Bucket.USERNAME]

    tocat = pd.read_csv(Settings.HC_PHONES_CSV, dtype=str, index_col=0)
    tocat[ExtraCols.PASSWORD.value] = np.nan
    hc_ec2 = pd.concat([hc_ec2, tocat], ignore_index=True)

    # to_drop = hc_ec2[hc_ec2['username'].isin(IgnoreUsers.HC)].index
    # hc_ec2 = hc_ec2.drop(to_drop)
  
    # Sort the dataframe such that rows with passwords come first
    # hc_ec2 = hc_ec2.sort_values(by=ExtraCols.PASSWORD.value, ascending=False)
    # Drop duplicates, keeping the first occurrence (which will have the password if available)
    hc_ec2 = hc_ec2.drop_duplicates(subset=ExtraCols.USER_PHONE.value, keep='last', ignore_index=True)
    # hc_ec2 = hc_ec2.reset_index(drop=True)
    hc_ec2.to_csv(Settings.HC_EC2_CSV)



def users_data():
    print("\nDownloading credentials from ec2...")
    os_type = get_os()
    if os.environ.get('mode', None) == 'cloud':
        pass
    if os.environ.get('mode', None) != 'testing':
        if os_type == "Linux" or os_type == "Darwin":
            run_shell_command("chmod +x src/download_users.sh")
            run_shell_command("/src/download_users.sh")
        elif os_type == "Windows":
            # Set-ExecutionPolicy RemoteSigned -Scope CurrentUser
            run_windows_shell_command("./download_users_windows.ps1")
        else:
            print(f"Unsupported OS: {os_type}")
            raise SystemExit
    
    combine_yahav_ec2()
    healthy_ec2()



def list_bucket(save_list=False) -> List[str]:
    # AWS_ACCESS_KEY_ID = Keys.AWS_ACCESS_KEY_ID
    # AWS_SECRET_ACCESS_KEY = Keys.AWS_SECRET_ACCESS_KEY

    # Set AWS credentials in environment
    # os.environ['AWS_ACCESS_KEY_ID'] = AWS_ACCESS_KEY_ID
    # os.environ['AWS_SECRET_ACCESS_KEY'] = AWS_SECRET_ACCESS_KEY

    # Create S3 client
    # s3 = boto3.client('s3', aws_access_key_id=AWS_ACCESS_KEY_ID, aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
    client = storage.Client()
    bucket = client.bucket(Settings.BUCKET_NAME)
    blobs = bucket.list_blobs()

    # response = s3.list_blobs(bucket_or_name=Settings.BUCKET_NAME)

    # If 'KeyCount' exists, use it to calculate total chunks
    # total_files = response.get('KeyCount', 0)
    total_files = len(list(blobs))
    total_chunks = (total_files // 1000) + 1 if total_files else 0

    blobs = bucket.list_blobs()
    files = []
    # Apply tqdm to the chunked retrieval (while loop)
    with tqdm(total=total_chunks, desc="Listing files chunks from GCS bucket") as pbar:
        # while True:
        #     for content in response.get('Contents', []):
        #         files.append(content['Key'])
        #
        #     # Check if there are more files to list
        #     if response.get('IsTruncated'):  # True if there are more files
        #         continuation_token = response.get('NextContinuationToken')
        #         response = s3.list_blobs(bucket_or_name=Settings.BUCKET_NAME, ContinuationToken=continuation_token)
        #         pbar.update(1)  # Update the progress bar after each chunk
        #     else:
        #         break
        for page in blobs.pages:
            page_blobs = list(page)  # materialize the page results
            for blob in page_blobs:
                files.append(blob.name)
            # Update progress bar after processing each page
            pbar.update(1)

    if save_list:
        list_to_csv(files, 'files.csv')
    return files




def get_bucket(skip=True) -> None:
    #TODO add feedback.csv files to the bucket!
    '''
    1.  if bucket exists, 
            open bucket. 
        else, create it and processed_files will be [].
    2.  list unprocessed files from bucket
    3.  add new files to bucket
    4.  extract all attributes from the new files
    5.  save to bucket
    '''
    if (not exists(Settings.BUCKET_CSV)) or (skip==False):
        bucket = pd.DataFrame(columns=Bucket.values())
        processed_filekeys = []
    else:
        bucket = pd.read_csv(Settings.BUCKET_CSV, dtype=str)
        processed_filekeys = bucket['filekey'].to_list()       
    
    filekeys = list_bucket()
    filekeys = [f for f in filekeys if f not in processed_filekeys]

    dfs = []
    for filekey in tqdm(filekeys, desc="Adding new files to database"):
        df = pd.DataFrame(columns=Bucket.values())
        df.loc[0, Bucket.FILEKEY] = filekey
        filename = filekey.split('/')[-1]
        for pattern in Patterns.values():
            if re.match(pattern.value, filename):
                username = filekey.split('/')[0]
                df.loc[0, Bucket.USERNAME] = username
                if username.startswith("hc_"):
                        df.loc[0, Bucket.ENTITY] = Entity.HC  # Ataxia will be resolved later
                elif len(username)==40:
                    df.loc[0, Bucket.ENTITY] = Entity.PD
                else:
                    df.loc[0, Bucket.ENTITY] = Entity.SA
                
                df.loc[0, Bucket.PATTERN] = pattern.name
                df.loc[0, Bucket.EXERCISE] = pattern.name.lower()
                if pattern.name not in ["REGISTRATION0", 'APKINSON']:
                    df.loc[0, Bucket.DATE] = extract_from_filename(filekey, 'date')
                    df.loc[0, Bucket.TIME] = extract_from_filename(filekey, 'time')
                    df.loc[0, Bucket.DATETIME] = extract_from_filename(filekey, 'datetime')

                if pattern.name!="UPDATE":
                    df.loc[0, Bucket.LANG] = extract_from_filename(filekey, 'language')
                
                if pattern.name in ['RECORDING', 'RECORDING1', 'FOG', 'SDQ', 'WOQ', 'UPDATE']:
                    df.loc[0, Bucket.SESSION] = filekey.split('/')[1]
                
                if pattern.name in ['UPDRS', 'UPDRS3', 'UPDRS124']:
                    df.loc[0, 'timing'] = extract_from_filename(filekey, 'timing')                    

                if pattern.name=='RECORDING':
                    df.loc[0, Bucket.EXERCISE] = extract_from_filename(filekey, 'exercise') # override
                    df.loc[0, Bucket.TIMING] = extract_from_filename(filekey, 'timing')
                    df.loc[0, Bucket.ONMED] = extract_from_filename(filekey, 'onmed')
                    df.loc[0, Bucket.ONOFF] = extract_from_filename(filekey, 'onoff')
                elif pattern.name=='RECORDING1':
                    df.loc[0, Bucket.EXERCISE] = extract_from_filename(filekey, 'exercise') # override
                    df.loc[0, Bucket.TIMING] = extract_from_filename(filekey, 'timing')
                    df.loc[0, Bucket.ONMED] = OnMed.ONMED if filekey.endswith("_on") else OnMed.NOTONMED
                    df.loc[0, Bucket.ONOFF] = FeelOnOff.UNKNOWN
                
                dfs.append(df)
                break
    if dfs:
        dfs = pd.concat(dfs, ignore_index=True)
        bucket = pd.concat([bucket, dfs], ignore_index=True)
        bucket.to_csv(Settings.BUCKET_CSV, index=False)
    
    

def download_csv_to_df(filekey: str, bucket_name = Settings.BUCKET_NAME, s3 = None) -> pd.DataFrame:
    if not s3:
        s3 = get_s3()
    ENCRYPTION_KEY = Keys.ENCRYPTION_KEY
    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete=False) as temp_file:
        temp_file_name = temp_file.name
    
    try:
        # Download the file from S3 to the temporary location
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(filekey)
        blob.download_to_filename(temp_file_name)
        fernet = Fernet(ENCRYPTION_KEY)
        with open(temp_file_name, 'rb') as file:
            encrypted_data = file.read()
        decrypted_data = fernet.decrypt(encrypted_data)
        with open(temp_file_name, 'wb') as file:
            file.write(decrypted_data)
        # Read the CSV file and convert it to a df
        df = pd.read_csv(temp_file_name, dtype=str)
        return df
    finally:
        # Clean up the temporary file
        os.remove(temp_file_name)



def get_raw_data(skip=True, print_filekey=False):
    '''
    1. open the bucket file
    2. for each file key that is registration or qnnr, extract the relevant files;
    3. also, append  each one of them to the relevant csv.
    4. C'est tout!
    '''
    pd.set_option('future.no_silent_downcasting', True)

    def check_exist_and_return(filepath: str, skip=skip) -> pd.DataFrame:
        if (not exists(filepath)) or (skip==False):
            return pd.DataFrame()
        else:
            return pd.read_csv(filepath, dtype=str)
    
    def put_filekey_first(df: pd.DataFrame) -> pd.DataFrame:
        if 'filekey' in df.columns:
            cols = df.columns.tolist()
            cols.remove('filekey')
            df = df[['filekey'] + cols]
        else:
            print("The DataFrame does not contain a 'filekey' column.")
        return df

    bucket = pd.read_csv(Settings.BUCKET_CSV, dtype=str)
    raw = pd.read_csv(Settings.RAW_CSV, dtype=str) if exists(Settings.RAW_CSV) else pd.DataFrame()
    if exists(Settings.RAW_CSV):
        new_filekeys = bucket.loc[~bucket[Bucket.FILEKEY].isin(raw[Bucket.FILEKEY])].copy()
    else:
        new_filekeys = bucket.copy()
    

    updrs = check_exist_and_return(Settings.UPDRS_CSV)
    moca = check_exist_and_return(Settings.MOCA_CSV)
    pdq8 = check_exist_and_return(Settings.PDQ8_CSV)
    fog = check_exist_and_return(Settings.FOG_CSV)
    sdq = check_exist_and_return(Settings.SDQ_CSV)
    woq = check_exist_and_return(Settings.WOQ_CSV)
    registration = check_exist_and_return(Settings.REGISTRATION_CSV)
    update = check_exist_and_return(Settings.UPDATE_CSV)
    medications = check_exist_and_return(Settings.MEDICATION_CSV)

    for ii,row in tqdm(new_filekeys.iterrows(), desc="Extracting raw data", total=len(new_filekeys)):
        filekey = row[Bucket.FILEKEY]
        pattern = row[Bucket.PATTERN]
        if print_filekey:
            print(filekey)
        if pattern in ["UPDRS", "UPDRS3", "UPDRS124"]:
            if Qnnrs.UPDRS1.value not in row or pd.isna(row[Qnnrs.UPDRS1.value]) or row[Qnnrs.UPDRS1.value]=='':
                df = download_csv_to_df(filekey)
                new_filekeys.loc[ii, Qnnrs.UPDRS1] = df.loc[0, UPDRS.updrs1.value].astype(int).sum()
                new_filekeys.loc[ii, Qnnrs.UPDRS2] = df.loc[0, UPDRS.updrs2.value].astype(int).sum()
                new_filekeys.loc[ii, Qnnrs.UPDRS3] = df.loc[0, UPDRS.updrs3.value].astype(int).sum()
                new_filekeys.loc[ii, Qnnrs.UPDRS4] = df.loc[0, UPDRS.updrs4.value].astype(int).sum()
                new_filekeys.loc[ii, Qnnrs.HY] = df.loc[0, UPDRS.hy.value]
                new_filekeys.loc[ii, Bucket.SAMPLER] = df.loc[0, Bucket.SAMPLER] if Bucket.SAMPLER in df else pd.NaT
                if updrs.empty or (filekey not in updrs[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    updrs = pd.concat([updrs, df], ignore_index=True)
        
        if pattern=="MOCA":
            if Qnnrs.MOCA.value not in row or pd.isna(row[Qnnrs.MOCA.value]) or row[Qnnrs.MOCA.value]=='':
                df = download_csv_to_df(filekey)
                df = df.replace({"True": 1, "False": 0})
                new_filekeys.loc[ii, Qnnrs.MOCA] = df.loc[0, MoCA.moca.value].astype(int).sum()
                new_filekeys.loc[ii, Bucket.SAMPLER] = df.loc[0, Bucket.SAMPLER] if Bucket.SAMPLER in df else pd.NaT
                if moca.empty or (filekey not in moca[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    moca = pd.concat([moca, df], ignore_index=True)
        
        if pattern=="PDQ8":
            if Qnnrs.PDQ8.value not in row or pd.isna(row[Qnnrs.PDQ8.value]) or row[Qnnrs.PDQ8.value]=='':
                df = download_csv_to_df(filekey)
                new_filekeys.loc[ii, Qnnrs.PDQ8] = df.loc[0, PDQ8.pdq8.value].astype(int).sum()
                new_filekeys.loc[ii, Bucket.SAMPLER] = df.loc[0, Bucket.SAMPLER] if Bucket.SAMPLER in df else pd.NaT
                if pdq8.empty or (filekey not in pdq8[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    pdq8 = pd.concat([pdq8, df], ignore_index=True)
        
        if pattern=="FOG":
            if Qnnrs.FOG.value not in row or pd.isna(row[Qnnrs.FOG.value]) or row[Qnnrs.FOG.value]=='':
                df = download_csv_to_df(filekey)
                new_filekeys.loc[ii, Qnnrs.FOG] = df.loc[0, FOG.fog.value].astype(int).sum()
                if fog.empty or (filekey not in fog[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    fog = pd.concat([fog, df], ignore_index=True)
        
        if pattern=="SDQ":
            if Qnnrs.SDQ.value not in row or pd.isna(row[Qnnrs.SDQ.value]) or row[Qnnrs.SDQ.value]=='':
                df = download_csv_to_df(filekey)
                df = df.replace({"True": 1, "False": 0})
                score = df.loc[0, SDQ.sdq.value[:-1]].astype(int).sum()
                respiratory = 2.5 if df.loc[0, SDQ.sdq.value[-1]]=="True" else 0.5
                score += respiratory
                new_filekeys.loc[ii, Qnnrs.SDQ] = score
                if sdq.empty or (filekey not in sdq[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    sdq = pd.concat([sdq, df], ignore_index=True)
        
        if pattern=="WOQ":
            if Qnnrs.WOQ_PRE.value not in row or pd.isna(row[Qnnrs.WOQ_PRE.value]) or row[Qnnrs.WOQ_PRE.value]=='':
                df = download_csv_to_df(filekey)
                df = df.replace({"True": 1, "False": 0})
                try:
                    new_filekeys.loc[ii, Qnnrs.WOQ_PRE] = df.loc[0, WOQ.pre.value].astype(int).sum()
                    new_filekeys.loc[ii, Qnnrs.WOQ_POST] = df.loc[0, WOQ.pre.value].astype(int).sum() - df.loc[0, WOQ.post.value].astype(int).sum()
                except:
                    new_filekeys.loc[ii, Qnnrs.WOQ_PRE] = -1
                    new_filekeys.loc[ii, Qnnrs.WOQ_POST] = -1
                if woq.empty or (filekey not in woq[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    woq = pd.concat([woq, df], ignore_index=True)

        if pattern=="REGISTRATION":
            if Registration.BIRTHDATE.value not in row or pd.isna(row[Registration.BIRTHDATE.value]) or row[Registration.BIRTHDATE.value]=='':
                df = download_csv_to_df(filekey)
                for col in Registration.values():
                    if col in df:
                        new_filekeys.loc[ii, col] = str(df.loc[0, col])
                if registration.empty or (filekey not in registration[Bucket.FILEKEY].tolist()):
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    registration = pd.concat([registration, df], ignore_index=True)
    
        if pattern=="UPDATE":
            if Update.IS_DBS.value not in row or pd.isna(row[Update.IS_DBS.value]) or row[Update.IS_DBS.value]=='':
                df = download_csv_to_df(filekey)
                for col in Update.values():
                    if col in df:
                        new_filekeys.loc[ii, col] = str(df.loc[0, col])
                if update.empty or (filekey not in update[Bucket.FILEKEY].tolist()):
                    df = download_csv_to_df(filekey)
                    df[Bucket.FILEKEY] = filekey
                    df = put_filekey_first(df)
                    update = pd.concat([update, df], ignore_index=True)

        if pattern=="MEDICATIONS":
            if medications.empty or (filekey not in medications[Bucket.FILEKEY].tolist()):
                df = download_csv_to_df(filekey)
                df[Bucket.FILEKEY] = filekey
                df = put_filekey_first(df)
                medications = pd.concat([medications, df.astype(str)], ignore_index=True)

        
    updrs.to_csv(Settings.UPDRS_CSV, index=False)
    moca.to_csv(Settings.MOCA_CSV, index=False)
    pdq8.to_csv(Settings.PDQ8_CSV, index=False)
    fog.to_csv(Settings.FOG_CSV, index=False)
    sdq.to_csv(Settings.SDQ_CSV, index=False)
    woq.to_csv(Settings.WOQ_CSV, index=False)
    update.to_csv(Settings.UPDATE_CSV, index=False)
    registration.to_csv(Settings.REGISTRATION_CSV, index=False)
    medications.to_csv(Settings.MEDICATION_CSV, index=False)

    raw = pd.concat([raw, new_filekeys], ignore_index=True)
    raw = raw.sort_values(by=['date', 'username', 'time'], ascending=False, ignore_index=True)
    raw.to_csv(Settings.RAW_CSV, index=False)



def patient_paradigm(session_df: pd.DataFrame, counter: bool = False) -> str:
    # Define the expected sections and their respective markers
    pre_marker = 'B'
    questionnaire_marker = 'Q'
    post_marker = 'A'
    missing_marker = '~'
    
    # Define the expected exercises and questionnaires
    pre_exercises = [ex for ex in Exercise]
    questionnaires = ['fog', 'sdq', 'woq', 'update']
    post_exercises = [ex for ex in Exercise]
    
    # Initialize the completeness string
    completeness = []

    # Check for pre files
    pre_files = session_df[(session_df['exercise'].isin(pre_exercises)) & (session_df['timing'] == 'pre')]
    for ex in pre_exercises:
        matches = pre_files[pre_files['exercise'] == ex]
        if not matches.empty:
            # Add the marker and the number of tags based on the counter setting
            tag = "'" * (len(matches) - 1) if counter and len(matches) > 1 else "'"
            completeness.append(f"{pre_marker}{tag}" if len(matches) > 1 else pre_marker)
        else:
            completeness.append(missing_marker)
    completeness.append('|')

    # Check for questionnaire files
    for q in questionnaires:
        matches = session_df[session_df['exercise'] == q]
        if not matches.empty:
            tag = "'" * (len(matches) - 1) if counter and len(matches) > 1 else "'"
            completeness.append(f"{questionnaire_marker}{tag}" if len(matches) > 1 else questionnaire_marker)
        else:
            completeness.append(missing_marker)
    completeness.append('|')

    # Check for post files
    post_files = session_df[(session_df['exercise'].isin(post_exercises)) & (session_df['timing'] == 'post')]
    for ex in post_exercises:
        matches = post_files[post_files['exercise'] == ex]
        if not matches.empty:
            tag = "'" * (len(matches) - 1) if counter and len(matches) > 1 else "'"
            completeness.append(f"{post_marker}{tag}" if len(matches) > 1 else post_marker)
        else:
            completeness.append(missing_marker)

    # Combine completeness markers into a single string
    completeness_str = '[' + ''.join(completeness) + ']'
    return completeness_str


def healthy_paradigm(session_df: pd.DataFrame, counter: bool = False) -> str:
    # Define the expected sections and their respective markers
    marker = 'H'
    missing_marker = '~'
    
    # Define the expected exercises and questionnaires
    exercises = [ex for ex in Exercise]
    
    # Initialize the completeness string
    completeness = []

    # Check for healthy session files
    h_files = session_df[(session_df['exercise'].isin(exercises)) & (session_df['timing'] == 'healthy')]
    for ex in exercises:
        matches = h_files[h_files['exercise'] == ex]
        if not matches.empty:
            count = len(matches)
            tag = "'" * (count - 1) if counter and count > 1 else "'"
            completeness.append(f"{marker}{tag}" if count > 1 else marker)
        else:
            completeness.append(missing_marker)

    # Combine completeness markers into a single string
    completeness_str = '[' + ''.join(completeness) + ']'
    return completeness_str


def sampler_paradigm(session_df: pd.DataFrame, counter: bool = False) -> str:
    # Define the missing marker
    missing_marker = '~'
    
    # Initialize the completeness list to hold results
    completeness = []

    # Iterate over the values in SamplerQnnrs
    for ex in SamplerQnnrs.values():
        # Filter the dataframe based on exercise and timing
        if ex['timing']:  # If timing is specified, filter both exercise and timing
            matches = session_df[(session_df['exercise'] == ex['exercise']) & (session_df['timing'] == ex['timing'])]
        else:  # If timing is not specified, only filter by exercise
            matches = session_df[session_df['exercise'] == ex['exercise']]

        # Check for matches and add the corresponding marker
        if not matches.empty:
            count = len(matches)
            tag = "'" * (count - 1) if counter and count > 1 else "'"
            marker = ex['marker']
            completeness.append(f"{marker}{tag}" if count > 1 else marker)
        else:
            completeness.append(missing_marker)

    # Combine completeness markers into a single string
    completeness_str = '[' + ','.join(completeness) + ']'
    return completeness_str


def add_sampler_phone(df):
    samplers = pd.read_csv(Settings.SAMPLERS_CSV, dtype=str, usecols=[Bucket.SAMPLER, ExtraCols.SAMPLER_USERNAME, ExtraCols.HEBREW])
    df[Bucket.SAMPLER] = df[Bucket.SAMPLER].apply(sampler_phone_to_name)
    df.rename(columns={Bucket.SAMPLER: ExtraCols.SAMPLER_USERNAME}, inplace=True)
    df = pd.merge(df, samplers, how='left', on=[ExtraCols.SAMPLER_USERNAME])
    return df


def add_patient_phone(df):
    phones_pd = pd.read_csv(Settings.USERS_EC2_CSV, dtype=str, usecols=[ExtraCols.USER_PHONE, Bucket.USERNAME, ExtraCols.PASSWORD])
    phones_hc = pd.read_csv(Settings.HC_EC2_CSV, dtype=str, usecols=[ExtraCols.USER_PHONE, Bucket.USERNAME, ExtraCols.PASSWORD])
    phones = pd.concat([phones_pd, phones_hc], ignore_index=True)
    df = pd.merge(df, phones, how='left', on=[Bucket.USERNAME])
    return df 


def add_caregiver_phone(df):
    sheba = pd.read_csv(Settings.SHEBA_DATABASE, dtype=str)
    sheba['caregiver'] = sheba['caregiver'].apply(standardize_phone_number)
    sheba['טלפון'] = sheba['טלפון'].apply(standardize_phone_number)
    sheba = sheba[['טלפון', 'caregiver']]
    sheba.columns = [ExtraCols.USER_PHONE, ExtraCols.CAREGIVER_PHONE]
    ichilov = pd.read_csv(Settings.ICHILOV_DATABASE, dtype=str)
    ichilov['caregiver'] = ichilov['caregiver'].apply(standardize_phone_number)
    ichilov['טלפון'] = ichilov['טלפון'].apply(standardize_phone_number)
    ichilov = ichilov[['טלפון', 'caregiver']]
    ichilov.columns = [ExtraCols.USER_PHONE, ExtraCols.CAREGIVER_PHONE]
    both_med_centers = pd.concat([sheba, ichilov])
    both_med_centers = both_med_centers[~pd.isna(both_med_centers[ExtraCols.CAREGIVER_PHONE])]
    df = pd.merge(df, both_med_centers, how='left', on=[ExtraCols.USER_PHONE])
    return df 


def add_session_number(df):
    df = df.sort_values(by=['username', 'datetime'])
    df['session_number'] = None  # Initialize column
    for username, user_df in df.groupby('username'):
        user_df = user_df.sort_values(by='datetime')
        user_sessions = user_df.drop_duplicates(subset='session').sort_values(by='datetime')
        for i, session in enumerate(user_sessions['session'], start=1):
            df.loc[(df['username'] == username) & (df['session'] == session), 'session_number'] = i
    df['session_number'] = df['session_number'].astype('Int64')  # Ensure it's integer type
    return df


def add_session_to_all(df):
    rows_with_session = df[df['session'].notna()].copy()
    rows_without_session = df[df['session'].isna()].copy()

    def find_closest_session(row):
        same_user_sessions = rows_with_session[rows_with_session['username'] == row['username']]
        if not same_user_sessions.empty and pd.notna(row['datetime']):
            same_user_sessions = same_user_sessions.dropna(subset=['datetime'])
            if not same_user_sessions.empty:
                closest_row = same_user_sessions.iloc[(same_user_sessions['datetime'] - row['datetime']).abs().argsort()[:1]]
                return closest_row['session'].values[0]
        return pd.NA

    rows_without_session['session'] = rows_without_session.apply(find_closest_session, axis=1, result_type='expand')
    df = pd.concat([rows_with_session, rows_without_session], ignore_index=True)
    return df


def add_sampler_to_HC(df):
    for session in df['session'].unique():
        registration_file = df[(df['session'] == session) & (df['exercise'] == 'registration') & (df['entity'].isin(['HC', 'AX']))]
        if not registration_file.empty:
            sampler_username = registration_file['sampler_username'].values[0]
            df.loc[df['session'] == session, 'sampler_username'] = sampler_username
    return df


def remove_qnnrs_duplicates(df: pd.DataFrame):
    exercises_of_interest = ['fog', 'sdq', 'woq', 'update']
    filtered_data_for_duplicates = df[df['exercise'].isin(exercises_of_interest)]
    filtered_data_for_duplicates = filtered_data_for_duplicates.sort_values(by=['session', 'exercise', 'datetime'])
    filtered_data_no_duplicates = filtered_data_for_duplicates.drop_duplicates(subset=['session', 'exercise'], keep='first')

    remaining_data = df.drop(index=filtered_data_for_duplicates.index.difference(filtered_data_no_duplicates.index))
    remaining_data = remaining_data.reset_index(drop=True)
    return remaining_data


def propagate_values(df: pd.DataFrame) -> pd.DataFrame:
    pd.set_option('future.no_silent_downcasting', True)
    '''
    'age', 'updrs', 'H&Y'
    '''
    df = df.sort_values(by=['username', 'datetime'])

    columns_to_fill = ['healthy_name', 'moca', 'pdq8', 'updrs1' ,'updrs2','updrs4', 
                       'fog', 'sdq', 'woq_pre', 'woq_post', 
                        'is_dbs', 'is_change_in_smoking_routine', 'smoking_routine', 
                        'is_paroxysmal', 'is_vocal_cords_damage', 'is_change_in_voice', 
                        'is_treated', 'is_practicing', 'is_satisfied', 'is_change_in_medication', 
                        'birth_date', 'gender', 'mother_tongue', 'year_of_diagnosis', 
                        'respiratory_disorders', 'dbs', 'medical_center', 'sleep_talk', 
                        'constipation', 'falling', 'smell', 'genetic']
        
    for (username, session), group in tqdm(df.groupby(['username', 'session']), desc="Propagating values"):
        for col in columns_to_fill:
            unique_values = group[col].dropna().unique()
            if len(unique_values) == 1:
                df.loc[(df['username'] == username) & (df['session'] == session), col] = unique_values[0]
            elif len(unique_values) > 1:
                Warning(f"Non-unique values for session {session} and column {col}!")
                df.loc[(df['username'] == username) & (df['session'] == session), col] = unique_values[0]

    df[columns_to_fill] = df.groupby('username')[columns_to_fill].ffill().infer_objects(copy=False)
    df[columns_to_fill] = df.groupby('username')[columns_to_fill].bfill().infer_objects(copy=False)
    df = df.sort_values(by=['date', 'username', 'time'], ascending=False)
    
    return df


def add_age(df: pd.DataFrame) -> pd.DataFrame:
    # Convert date columns to datetime and calculate age
    date = pd.to_datetime(df['date'], format=Settings.DATE_FORMAT, errors='coerce')
    bdate = pd.to_datetime(df['birth_date'], format=Settings.DATE_FORMAT, errors='coerce')
    df['age'] = ((date - bdate).dt.days / 365.25).astype(int)
    return df



def add_updrs_columns(df: pd.DataFrame) -> pd.DataFrame:

    # df = df.sort_values(by=['username', 'datetime'])

    # Create new columns for pre and post values
    df['updrs3_pre'] = None
    df['updrs3_post'] = None
    df['H&Y_pre'] = None
    df['H&Y_post'] = None

    # Propagate values within the same session first
    for session, session_data in df.groupby(['username', 'session']):
        pre_values = session_data[(session_data['exercise'] == 'updrs3') & (session_data['timing'] == 'pre')][['updrs3', 'H&Y']].dropna()
        post_values = session_data[(session_data['exercise'] == 'updrs3') & (session_data['timing'] == 'post')][['updrs3', 'H&Y']].dropna()

        if not pre_values.empty:
            df.loc[session_data.index, 'updrs3_pre'] = pre_values['updrs3'].values[0]
            df.loc[session_data.index, 'H&Y_pre'] = pre_values['H&Y'].values[0]

        if not post_values.empty:
            df.loc[session_data.index, 'updrs3_post'] = post_values['updrs3'].values[0]
            df.loc[session_data.index, 'H&Y_post'] = post_values['H&Y'].values[0]

    # Now propagate the values forward across sessions using forward-fill only for the specified columns
    df[['updrs3_pre', 'updrs3_post', 'H&Y_pre', 'H&Y_post']] = df.sort_values(by=['username', 'datetime'])\
                                                                .groupby('username')[['updrs3_pre', 'updrs3_post', 'H&Y_pre', 'H&Y_post']]\
                                                                .ffill()
    return df



def get_all_users(df: pd.DataFrame) -> pd.DataFrame:
    df = df[df['exercise']=='registration']
    df = df.drop(['filekey', 'exercise'], axis=1)
    df = df.reset_index(drop=True)
    return df




def get_all_files():
    print("Arranging data ...", end=' ')
    df = pd.read_csv(Settings.RAW_CSV, dtype=str)
    df = change_columns(df)
    df = resolve(df)

    df = add_sampler_phone(df)
    df = add_patient_phone(df)
    df = add_caregiver_phone(df)

    df['datetime'] = pd.to_datetime(df['datetime'], format=Settings.DATETIME, errors='coerce')
    df = add_session_to_all(df)

    real_sessions = df.copy()
    real_sessions = add_session_number(real_sessions)
    real_sessions = add_sampler_to_HC(real_sessions)

    df = resolve_sessions(df)
    df = add_session_number(df)
    df = remove_qnnrs_duplicates(df)
    df = add_sampler_to_HC(df)
    print("Done!")

    df = propagate_values(df)
    df = add_age(df)
    df = add_updrs_columns(df)
    df = df.sort_values(by=['date', 'username', 'time'], ascending=False)
    df.to_csv(Settings.ALL_FILES, index=False)
    
    sessions = get_sessions(df)
    sessions.to_csv(Settings.SESSIONS, index=False)
    real_sessions = get_sessions(real_sessions)
    real_sessions.to_csv(Settings.REAL_SESSIONS, index=False)

    all_users = get_all_users(df)
    # all_users.to_csv('../'+Settings.ALL_USERS, index=False)
    client = bigquery.Client()
    table_ref = f"vocaapp-440108.vocaapp_dataset.all_users"
    all_users["year_of_diagnosis"] = pd.to_datetime(
        all_users["year_of_diagnosis"], errors="coerce"
    ).dt.date  # Convert to DATE

    # Replace NaT with a default valid date or drop rows with NaT
    all_users["year_of_diagnosis"] = all_users["year_of_diagnosis"].fillna(pd.Timestamp("1900-01-01").date())

    # Convert other date/time fields as needed
    all_users["date"] = pd.to_datetime(all_users["date"], errors="coerce").dt.date
    all_users["time"] = pd.to_datetime(all_users["time"], errors="coerce").dt.time
    all_users["datetime"] = pd.to_datetime(all_users["datetime"], errors="coerce")

    # Clean invalid rows if necessary
    all_users = all_users.dropna(subset=["year_of_diagnosis", "date", "datetime"])
    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,  # Overwrite the table if it exists
        source_format=bigquery.SourceFormat.CSV,
    )
    job = client.load_table_from_dataframe(all_users, table_ref, job_config=job_config)

    job.result()





def get_caregiver_phone(sessions_df: pd.DataFrame) -> str:
    df = sessions_df.copy()
    # Convert time columns to datetime objects
    df['start'] = pd.to_datetime(df['start'], format=Settings.TIME_FORMAT)
    df['end'] = pd.to_datetime(df['end'], format=Settings.TIME_FORMAT)
    # df['registration'] = pd.to_datetime(df['registration'], format=Settings.TIME_FORMAT)

    # Iterate over each row where 'status' is 'PD'
    for pd_index, pd_row in df[df['entity'] == 'PD'].iterrows():
        # Filter rows where 'status' is 'HC' and 'sampler' matches
        matching_hcs = df[(df['entity'] == 'HC') &
                        # (df['exercise'] == 'registration') &
                        (df['sampler_username'] == pd_row['sampler_username']) &
                        (df['date'] == pd_row['date']) &
                        (df['start'] >= pd_row['start']) &
                        (df['start'] <= pd_row['end'])]
        # Check for errors or match
        if len(matching_hcs['username'].unique()) > 1:
            raise ValueError(f"Multiple matches found for PD at index {pd_index}")
        elif len(matching_hcs['username'].unique()) == 1:
            hc_username = matching_hcs.iloc[0]['username']
            hc_session = matching_hcs.iloc[0]['session']
            hc_full = matching_hcs.iloc[0]['user_paradigm']
            sessions_df.loc[pd_index, 'caregiver_username'] = hc_username
            sessions_df.loc[pd_index, 'caregiver_session'] = hc_session
            sessions_df.loc[pd_index, 'caregiver_paradigm'] = hc_full

    # match caregiver PHONE from recorded healthy
    healthy = pd.read_csv(Settings.USERS_CSV, dtype=str)
    filtered_healthy = healthy.dropna(subset=['user_phone'])
    healthy_dict = filtered_healthy.set_index('username')['user_phone'].to_dict()

    def replace_dict(x):
        if x in healthy_dict.keys():
            return healthy_dict[x]
        else:
            return None

    sessions_df['caregiver_phone'] = sessions_df['caregiver_username'].apply(lambda x: replace_dict(x))

    return sessions_df



def get_sessions(df: pd.DataFrame) -> pd.DataFrame:
    sessions_list = df['session'].dropna().unique().tolist()
    sessions = []
    for session in tqdm(sessions_list, desc="Acquiring sessions"):
        single_session = df[df.session==session].copy()
        single_session['datetime'] = pd.to_datetime(single_session['datetime'], format=Settings.DATETIME, errors='coerce')
        single_session = single_session.sort_values(by='datetime', ascending=True)
        session_df = pd.DataFrame(columns=['date', 'entity', 'username', 'session', 'session_number', 'start', 'end', 'duration',\
                                            'sampler_username', 'paradigm', 'user_phone', 'password', 'caregiver_phone'])

        session_df.loc[0, 'date'] = single_session['date'].tolist()[0]
        session_df.loc[0, 'username'] = single_session['username'].tolist()[0]
        
        entity = single_session['entity'].tolist()[0]
        session_df.loc[0, 'entity'] = entity
        session_df.loc[0, 'session'] = single_session['session'].tolist()[0]
        session_df.loc[0, 'session_number'] = single_session['session_number'].tolist()[0]
        session_df.loc[0, 'caregiver_phone'] = single_session['caregiver_phone'].tolist()[0]
        session_df.loc[0, 'user_phone'] = single_session['user_phone'].tolist()[0]
        session_df.loc[0, 'password'] = single_session['password'].tolist()[0]
        
        start_time = single_session['time'].tolist()[0]
        end_time = single_session['time'].tolist()[-1]
        session_df.loc[0, 'start'] = start_time
        session_df.loc[0, 'end'] = end_time
        duration = datetime.strptime(end_time, Settings.TIME_FORMAT) - datetime.strptime(start_time, Settings.TIME_FORMAT)
        total_seconds = duration.total_seconds()
        session_df.loc[0, 'duration'] = str(timedelta(seconds=total_seconds))
        
        sampler = single_session['sampler_username'].tolist()[0]
        session_df.loc[0, 'sampler_username'] = sampler
        # if isinstance(sampler, str) & (entity=='PD'):
        #     session_df.loc[0, 'paradigm_'] = sampler_paradigm(single_session)
        if entity == 'PD':
            user_paradigm = patient_paradigm(single_session)
            if 'sampler_username' in single_session.columns and pd.notna(single_session['sampler_username'].iloc[0]):
                sampler_paradigm_str = sampler_paradigm(single_session)
                user_paradigm += f"\n{sampler_paradigm_str}"
            session_df.loc[0, 'paradigm'] = user_paradigm
        elif entity == "HC":
            session_df.loc[0, 'paradigm'] = healthy_paradigm(single_session)
        elif entity == "AX":
            session_df.loc[0, 'paradigm'] = healthy_paradigm(single_session).replace("H","X")
        sessions.append(session_df)
    sessions = pd.concat(sessions, ignore_index=True)
    sessions = sessions.sort_values(by=['date', 'username', 'start'], ascending=False, ignore_index=True)
    # sessions = match_caregiver(sessions)

    
    # arrange columns and sort
    # sessions = sessions[['date', 'entity', 'username', 'session', 'number_of_session', 'start', 'end', 'duration',\
    #                                         'sampler_username', 'sampler_paradigm', 'user_paradigm', 'caregiver_username', 'caregiver_session',\
    #                                               'caregiver_paradigm', 'caregiver_phone', 'user_phone', 'password']]
    sessions = sessions.sort_values(by=['date', 'username', 'start'], ascending=False, ignore_index=True)
    return sessions



def color_sessions():
    sessions_df = pd.read_csv(Settings.SESSIONS, dtype=str)
    wb = Workbook()
    ws = wb.active

    # Define colors
    dark_purple = PatternFill(start_color=rgb_to_hex(Color.DARK_VV, 0.7), end_color=rgb_to_hex(Color.DARK_VV, 0.7), fill_type="solid")
    yellow = PatternFill(start_color=rgb_to_hex(Color.YELLOW, 0.7), end_color=rgb_to_hex(Color.YELLOW, 0.7), fill_type="solid")
    purple = PatternFill(start_color=rgb_to_hex(Color.VV, 0.7), end_color=rgb_to_hex(Color.VV, 0.7), fill_type="solid")
    light_green = PatternFill(start_color=rgb_to_hex(Color.GREEN, 0.7), end_color=rgb_to_hex(Color.GREEN, 0.7), fill_type="solid")

    # Add headers
    for col_num, column_title in enumerate(sessions_df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Append dataframe to worksheet
    for r_idx, row in sessions_df.iterrows():
        row_height = 30  # Default row height
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
            
            # # Enable text wrapping for the 'sampler questionnaires' column
            # if sessions_df.columns[c_idx] == 'sampler_paradigm':
            #     cell.alignment = Alignment(wrap_text=True)
            #     # Calculate the necessary row height
            #     lines = len(str(value).splitlines())
            #     row_height = max(row_height, lines * 15)  # Adjust the multiplier as necessary
            
            # Coloring cells based on conditions
            if sessions_df.columns[c_idx] == 'sampler_username' and pd.isnull(value):
                cell.fill = dark_purple
            elif sessions_df.columns[c_idx] in ['paradigm']:
                if not pd.isnull(value) and "~" in str(value):
                    cell.fill = dark_purple
            elif sessions_df.columns[c_idx] == 'entity':
                if value == "HC":
                    cell.fill = yellow
                elif value == "PD":
                    cell.fill = purple
            elif sessions_df.columns[c_idx] == 'paradigm' and not pd.isnull(value) and "~" in str(value):
                cell.fill = dark_purple

        ws.row_dimensions[r_idx + 2].height = row_height

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(Settings.COLORED_SESSIONS)



    
    



def main(): #TODO Make as entry point of function
    print('start')
    # get_bucket(skip=False)
    # users_data()
    # get_raw_data()
    get_all_files()
    # color_sessions()


    print("Completed!")


if __name__ == "__main__":
    main()
   


